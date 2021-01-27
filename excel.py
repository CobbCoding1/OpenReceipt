from openpyxl import Workbook, load_workbook
import os

# Add a receipt to the file
def addReceipt(receipt):
    # Open excel file

    file = open('filename.txt', 'r')
    excelFile = file.readline()
    file.close()
    print(excelFile)

    # Load workbook from file
    workbook = load_workbook(excelFile)
    sheet = workbook.active

    num = 0

    # Get the total occupied rows in the sheet
    for row in sheet['A']:
        num += 1

    # Add the values to the sheet, adding 1 to the number to get the first blank column
    sheet['A' + str(num+1)] = receipt[0]
    sheet['B' + str(num+1)] = receipt[1]
    sheet['C' + str(num+1)] = receipt[2]

    # Save file
    workbook.save(filename=excelFile)

    # Return sheet information
    return(sheet['A:C'])

def open_file(filename):
    # Open existing workbook
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Save workbook as the selected filename
    workbook.save(filename=filename)

    # Open and write filename to file
    file = open('filename.txt', 'w')
    file.write(filename)

    print('Wrote to file!')

def new_file(dirname):
    # Create new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Add workbook template
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Total'
    sheet['C1'] = 'Store'

    # Save the workbook
    workbook.save(filename=dirname + '/receipts.xlsx')

    # Select file as default excel file
    open_file(dirname + '/receipts.xlsx')
